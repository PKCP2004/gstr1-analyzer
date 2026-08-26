import io, re, zipfile
from pathlib import Path
from datetime import datetime
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Excel template mapping
# -----------------------------------------------------------------------------
# Each output group is mapped to the corresponding filed GSTR-1 section.
# IMPORTANT: The PDF's "Value" column is treated as TAXABLE VALUE.
# Invoice Value is always calculated as:
#     Taxable Value + IGST + CGST + SGST + Cess
#
# The output template does not expose every tax column for every section, so
# the source tax fields are retained internally and only the template headers
# are written to Excel.

GROUP_CONFIG = {
    '4A, 4B, 6C - B2B, DE Invoices': {
        'type': 'aggregate',
        'sections': [
            ('4A - Taxable outward supplies made to registered persons', 'full', 'Total'),
            ('4B - Taxable outward supplies made to registered persons attracting tax on reverse charge', 'full', 'Total'),
            ('6C - Deemed Exports', 'full', 'Total'),
        ],
    },
    '5A, 5B - B2C (Large) Invoices': {
        'heading': '5 - Taxable outward inter-state supplies made to unregistered persons',
        'layout': 'value_igst_cess',
    },
    '6A - Exports Invoices': {
        'heading': '6A – Exports (with/without payment)',
        'heading_alt': '6A - Exports (with/without payment)',
        'layout': 'value_igst_cess',
    },
    '6B- SEZ WOP': {
        'heading': '6B - Supplies made to SEZ unit or SEZ developer',
        'row_label': '- SEZWOP',
        'layout': 'value_igst_cess',
    },
    '6B- SEZ WP': {
        'heading': '6B - Supplies made to SEZ unit or SEZ developer',
        'row_label': '- SEZWP',
        'layout': 'value_igst',
    },
    '7 - B2C (Others)': {
        'heading': '7- Taxable supplies (Net of debit and credit notes) to unregistered persons',
        'heading_alt': '7- Taxable supplies',
        'layout': 'full',
        # Table 7 has an extra "Document Type" text column (e.g. "Net Value")
        # sitting between the record count and the ₹ figures, and a rate-wise
        # breakdown (5%, 12%, 18%...) above the real total. Anchor precisely
        # on the row that starts with "Total" followed by a record count,
        # rather than any line merely containing the word "Total" somewhere,
        # so a rate-wise row or stray text can't be picked up by mistake.
        'row_pattern': r'^Total\s+[0-9]',
    },
    '8 - Nil rated, exempted and non GST outward supplies': {'type': 'table8'},
    '9A - Amended B2B Invoices': {
        'type': 'aggregate',
        'sections': [
            ('9A - Amendment to taxable outward supplies made to registered person in returns of earlier tax periods in table 4 - B2B Regular', 'full', 'Amended amount - Total'),
            ('9A - Amendment to taxable outward supplies made to registered person in returns of earlier tax periods in table 4 - B2B Reverse charge', 'full', 'Amended amount - Total'),
            ('9A - Amendment to Deemed Exports in returns of earlier tax periods in table 6C', 'full', 'Amended amount - Total'),
        ],
    },
    '9A - Amended B2C (Large) Invoices': {
        'heading': '9A - Amendment to Inter-State supplies made to unregistered person',
        'layout': 'value_igst_cess',
        'row_pattern': 'Amended amount - Total',
    },
    '9A - Amended Exports Invoices': {
        'heading': '9A - Amendment to Export supplies in returns of earlier tax periods in table 6A',
        'layout': 'value_igst_cess',
        'row_pattern': 'Amended amount - Total',
    },
    '9B - Credit/Debit Notes (Registered)': {
        'heading': '9B - Credit/Debit Notes (Registered)',
        'layout': 'full',
        'row_pattern': 'Total - Net off debit/credit notes',
    },
    '9B - Credit / Debit Notes (Unregistered)': {
        'heading': '9B - Credit/Debit Notes (Unregistered)',
        'layout': 'value_igst_cess',
        'row_pattern': 'Total - Net off debit/credit notes',
    },
    '9C - Amended Credit/Debit Notes (Registered)': {
        'heading': '9C - Amended Credit/Debit Notes (Registered)',
        'layout': 'full',
        'row_pattern': 'Amended amount - Total',
    },
    '9C - Amended Credit/Debit Notes (Unregistered': {
        'heading': '9C - Amended Credit/Debit Notes (Unregistered)',
        'layout': 'value_igst_cess',
        'row_pattern': 'Amended amount - Total',
    },
    '10 - Amended B2C(Others)': {
        'heading': '10 - Amendment to taxable outward supplies made to unregistered person',
        'layout': 'full',
        'row_pattern': 'Amended amount - Total',
    },
    '11A - Amended Tax Liability (Advance Received)': {
        'heading': '11A - Amendment to advances received',
        'layout': 'full',
        'row_pattern': 'Amended amount - Total',
    },
    '11B - Amendment of Adjustment of Advances': {
        'heading': '11B - Amendment to advances adjusted',
        'layout': 'full',
        'row_pattern': 'Amended amount - Total',
    },
    '11A(1), 11A(2) - Tax Liability (Advances Received)': {
        'heading': '11A(1), 11A(2) - Advances received',
        'layout': 'full',
        'row_pattern': 'Total',
    },
    '11B(1), 11B(2) - Adjustment of Advances': {
        'heading': '11B(1), 11B(2) - Advance amount received in earlier tax period',
        'layout': 'full',
        'row_pattern': 'Total',
    },
    '12 - HSN-wise summary of outward supplies': {
        'heading': '12 - HSN-wise summary of outward supplies',
        'layout': 'full',
        'row_pattern': r'^Total\s+\d+\s+NA\s+',
        'hsn': True,
    },
}


def load_template(template_path):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    groups = []
    month_col = 2
    for c in range(2, ws.max_column + 1):
        title = ws.cell(3, c).value
        if title:
            # The "Month" header (column B) is a single label column, not a
            # data group with sub-headers. Treating it as a group was the
            # cause of the Month column being overwritten with 0 on every
            # row (its blank sub-header fell into the default "0" branch of
            # write_group_row right after the real month value was set).
            if str(title).strip().lower() == 'month':
                month_col = c
                continue
            end = c
            while end + 1 <= ws.max_column and ws.cell(3, end + 1).value is None:
                end += 1
            groups.append({
                'start': c,
                'end': end,
                'title': str(title),
                'headers': [str(ws.cell(4, x).value or '').strip() for x in range(c, end + 1)],
            })
    return wb, ws, groups, month_col


def norm(s):
    if s is None:
        return ''
    s = str(s)
    s = s.replace('\u2013', '-').replace('\u2014', '-').replace('\u2011', '-')
    s = s.replace('₹', ' ')
    return re.sub(r'[ \t]+', ' ', s).strip()


def row_numbers(line):
    line = norm(line)
    line = line.replace('I0', '0').replace('N0.00', '0.00').replace('F0', '0')
    return re.findall(r'(?<![A-Za-z])[-]?[0-9][0-9,]*(?:\.[0-9]+)?', line)


def num(s):
    try:
        s = str(s).replace(',', '').strip()
        return float(s) if s else 0.0
    except (TypeError, ValueError):
        return 0.0


def zero():
    return {
        'records': 0,
        'taxable': 0.0,
        'igst': 0.0,
        'cgst': 0.0,
        'sgst': 0.0,
        'cess': 0.0,
        'invoice': 0.0,
        'found': False,
        'source_line': '',
    }


def parse_row(line, layout='full', hsn=False):
    """Parse a single filed-summary row using the exact PDF column layout."""
    if not line:
        return zero()
    ns = row_numbers(line)
    if not ns:
        return zero()

    # The first number is the record count in all mapped summary rows.
    records = int(round(num(ns[0])))
    values = [num(x) for x in ns[1:]]
    d = zero()
    d['records'] = records
    d['found'] = True
    d['source_line'] = line

    # "Value" in the GSTR-1 summary is mapped as taxable value for this tool.
    d['taxable'] = values[0] if len(values) >= 1 else 0.0
    if layout == 'value_igst_cess':
        d['igst'] = values[1] if len(values) >= 2 else 0.0
        d['cess'] = values[2] if len(values) >= 3 else 0.0
    elif layout == 'value_igst':
        d['igst'] = values[1] if len(values) >= 2 else 0.0
    else:  # value + IGST + CGST + SGST + Cess
        d['igst'] = values[1] if len(values) >= 2 else 0.0
        d['cgst'] = values[2] if len(values) >= 3 else 0.0
        d['sgst'] = values[3] if len(values) >= 4 else 0.0
        d['cess'] = values[4] if len(values) >= 5 else 0.0

    d['invoice'] = d['taxable'] + d['igst'] + d['cgst'] + d['sgst'] + d['cess']
    return d


def find_heading(lines, heading, heading_alt=None):
    targets = [norm(heading).lower()]
    if heading_alt:
        targets.append(norm(heading_alt).lower())
    for i, line in enumerate(lines):
        low = line.lower()
        if any(t in low for t in targets):
            return i
    return None


def find_row_after_heading(lines, heading, row_pattern=None, heading_alt=None, row_label=None, max_window=12):
    """Find a data row only inside the relevant section window."""
    idx = find_heading(lines, heading, heading_alt)
    if idx is None:
        return None

    start = idx + 1
    end = min(len(lines), start + max_window)
    window = lines[start:end]

    if row_label:
        for line in window:
            if norm(row_label).lower() in line.lower():
                return line

    if row_pattern:
        pattern = re.compile(row_pattern, re.I) if not isinstance(row_pattern, re.Pattern) else row_pattern
        for line in window:
            if pattern.search(line):
                return line

    # Default: first summary Total row. Avoid differential rows where possible.
    for line in window:
        low = line.lower()
        if ('net differential amount' in low or 'net differential' in low) and 'total' not in low.split('net differential')[0]:
            continue
        if re.search(r'\bTotal\b|\bNet Total\b|Amended amount', line, re.I):
            return line
    return None


def aggregate(parsed_rows):
    d = zero()
    found = False
    sources = []
    for p in parsed_rows:
        if not p.get('found'):
            continue
        found = True
        for k in ['records', 'taxable', 'igst', 'cgst', 'sgst', 'cess', 'invoice']:
            d[k] += p.get(k, 0) or 0
        if p.get('source_line'):
            sources.append(p['source_line'])
    d['found'] = found
    d['source_line'] = ' || '.join(sources)
    return d


def extract_pdf(pdf_bytes):
    lines = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            raw = page.extract_text() or ''
            for raw_line in raw.splitlines():
                line = norm(raw_line)
                if line:
                    lines.append(line)

    text = '\n'.join(lines)
    fy = re.search(r'Financial year\s+(\d{4}-\d{2})', text, re.I)
    tp = re.search(r'Tax period\s+([A-Za-z]+)', text, re.I)
    gstin = re.search(r'1 GSTIN\s+([0-9A-Z]{15})', text, re.I)
    arn = re.search(r'\(c\) ARN\s+([A-Z0-9]+)', text, re.I)
    arn_date = re.search(r'\(d\) ARN date\s+(\d{2}/\d{2}/\d{4})', text, re.I)

    month = None
    if fy and tp:
        try:
            month_num = datetime.strptime(tp.group(1)[:3], '%b').month
            start_year = int(fy.group(1)[:4])
            year = start_year if month_num >= 4 else start_year + 1
            month = datetime(year, month_num, 1)
        except Exception:
            month = None

    # IMPORTANT: Do NOT use a plain '\bNIL\b' search here. Normal GSTR-1
    # filed PDFs contain the words 'Nil rated' / 'Nil' in Table 8 even when
    # the return is NOT a NIL return. That old check therefore marked every
    # PDF as NIL.
    #
    # We first look for an explicit NIL-return marker. If the filed PDF has no
    # explicit marker, we later classify it as NIL only when every mapped
    # section is genuinely zero (no records and no values).
    filename_hint = ''
    explicit_nil = bool(re.search(
        r'(?:NIL\s+RETURN|RETURN\s+TYPE\s*[:\-]?\s*NIL|FILING\s+TYPE\s*[:\-]?\s*NIL|STATUS\s*[:\-]?\s*NIL)',
        text,
        re.I,
    ))
    is_nil = explicit_nil

    data = {
        'fy': fy.group(1) if fy else None,
        'tax_period': tp.group(1) if tp else None,
        'gstin': gstin.group(1) if gstin else None,
        'arn': arn.group(1) if arn else None,
        'arn_date': arn_date.group(1) if arn_date else None,
        'month': month,
        'is_nil': is_nil,
        'filing_status': 'NIL FILED' if is_nil else 'FILED',
        'sections': {},
        'audit': [],
    }

    # --- Section mappings ----------------------------------------------------
    for group, cfg in GROUP_CONFIG.items():
        if cfg.get('type') == 'table8':
            values = {}
            idx = find_heading(lines, '8 - Nil rated, exempted and non GST outward supplies')
            if idx is not None:
                for label, key in [('- Nil', 'nil'), ('- Exempted', 'exempted'), ('- Non-GST', 'non_gst')]:
                    found_line = None
                    for line in lines[idx + 1:min(len(lines), idx + 10)]:
                        if line.lower().startswith(label.lower()):
                            found_line = line
                            ns = row_numbers(line)
                            values[key] = num(ns[-1]) if ns else 0.0
                            break
                    if found_line:
                        data['audit'].append((group, found_line, 'special Table 8 value'))
            data['sections'][group] = {
                'nil': values.get('nil', 0.0),
                'exempted': values.get('exempted', 0.0),
                'non_gst': values.get('non_gst', 0.0),
                'found': idx is not None,
            }
            continue

        if cfg.get('type') == 'aggregate':
            parts = []
            for heading, layout, row_pattern in cfg['sections']:
                line = find_row_after_heading(lines, heading, row_pattern=row_pattern)
                parsed = parse_row(line, layout=layout) if line else zero()
                parts.append(parsed)
                if line:
                    data['audit'].append((group, line, heading))
            data['sections'][group] = aggregate(parts)
            continue

        heading = cfg.get('heading')
        line = find_row_after_heading(
            lines,
            heading,
            row_pattern=cfg.get('row_pattern'),
            heading_alt=cfg.get('heading_alt'),
            row_label=cfg.get('row_label'),
            max_window=15,
        )
        parsed = parse_row(line, layout=cfg.get('layout', 'full'), hsn=cfg.get('hsn', False)) if line else zero()
        data['sections'][group] = parsed
        if line:
            data['audit'].append((group, line, heading))

    # If the PDF does not contain an explicit NIL marker, classify it as a
    # NIL filing only when the extracted GSTR-1 summary is completely zero.
    # This is safe for a genuine NIL GSTR-1 and, crucially, avoids treating
    # the ordinary 'Nil rated' wording in Table 8 as a NIL-return marker.
    if not is_nil:
        has_records = False
        has_values = False
        for section in data['sections'].values():
            if not isinstance(section, dict):
                continue
            if (section.get('records', 0) or 0) != 0:
                has_records = True
                break
            for key in ('taxable', 'igst', 'cgst', 'sgst', 'cess', 'invoice', 'nil', 'exempted', 'non_gst'):
                if abs(float(section.get(key, 0) or 0)) > 0.000001:
                    has_values = True
                    break
            if has_values:
                break
        if not has_records and not has_values:
            is_nil = True

    data['is_nil'] = is_nil
    data['filing_status'] = 'NIL FILED' if is_nil else 'FILED'

    # Filed-copy reconciliation: Table 6A total should agree with the final
    # outward-supply liability when no other taxable outward tables contain
    # values. Keep this as an audit flag rather than changing extracted data.
    total_liability = None
    for line in lines:
        m = re.search(r'Total Liability \(Outward supplies other than Reverse charge\)\s+([-]?[0-9,]+(?:\.[0-9]+)?)', line, re.I)
        if m:
            total_liability = num(m.group(1))
            break
    data['total_liability'] = total_liability
    six_a = data['sections'].get('6A - Exports Invoices', zero())
    data['reconciliation'] = {
        'total_liability': total_liability,
        '6A_invoice_value': six_a.get('invoice', 0.0),
        'difference': (total_liability - six_a.get('invoice', 0.0)) if total_liability is not None else None,
    }

    # For a NIL-filed month, explicitly zero every mapped section. This makes
    # the month appear in the same report with zero values instead of being
    # mistaken for a missing month.
    if is_nil:
        for key in list(data['sections']):
            data['sections'][key] = zero()
        data['total_liability'] = 0.0
        data['reconciliation'] = {
            'total_liability': 0.0,
            '6A_invoice_value': 0.0,
            'difference': 0.0,
        }

    return data


def copy_row_style(ws, source_row, target_row):
    # NOTE: openpyxl's cell._style is a mutable array shared by reference,
    # not copied by value. Assigning `dst._style = src._style` directly (as
    # earlier versions of this helper did) makes the two cells point at the
    # SAME underlying style object, so a later change to one row's format
    # (e.g. the Total row) silently mutates every other row that ever
    # shared that reference. Copying each style component individually
    # avoids that bleed-through.
    for c in range(1, ws.max_column + 1):
        src = ws.cell(source_row, c)
        dst = ws.cell(target_row, c)
        if src.has_style:
            if src.font:
                dst.font = src.font.copy()
            if src.border:
                dst.border = src.border.copy()
            if src.fill:
                dst.fill = src.fill.copy()
            if src.alignment:
                dst.alignment = src.alignment.copy()
            if src.protection:
                dst.protection = src.protection.copy()
        if src.number_format:
            dst.number_format = src.number_format
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def write_group_row(ws, row, group, d):
    headers = group['headers']
    for i, header in enumerate(headers):
        c = group['start'] + i
        h = header.strip().lower()
        if h.startswith('no. of records'):
            value = d.get('records', 0)
        elif h == 'invoice value':
            # ALWAYS use all source tax components, even if some tax columns
            # are not exposed in this particular output group.
            value = d.get('invoice', 0)
        elif h == 'taxable value':
            value = d.get('taxable', 0)
        elif h == 'igst':
            value = d.get('igst', 0)
        elif h == 'cgst':
            value = d.get('cgst', 0)
        elif h == 'sgst':
            value = d.get('sgst', 0)
        elif h == 'cess':
            value = d.get('cess', 0)
        elif h == 'nil':
            value = d.get('nil', 0)
        elif h == 'exempted':
            value = d.get('exempted', 0)
        elif h == 'non-gst':
            value = d.get('non_gst', 0)
        else:
            value = 0
        ws.cell(row, c).value = value


STATE_CODES = {
    '01': 'Jammu & Kashmir', '02': 'Himachal Pradesh', '03': 'Punjab',
    '04': 'Chandigarh', '05': 'Uttarakhand', '06': 'Haryana',
    '07': 'Delhi', '08': 'Rajasthan', '09': 'Uttar Pradesh',
    '10': 'Bihar', '11': 'Sikkim', '12': 'Arunachal Pradesh',
    '13': 'Nagaland', '14': 'Manipur', '15': 'Mizoram',
    '16': 'Tripura', '17': 'Meghalaya', '18': 'Assam',
    '19': 'West Bengal', '20': 'Jharkhand', '21': 'Odisha',
    '22': 'Chhattisgarh', '23': 'Madhya Pradesh', '24': 'Gujarat',
    '25': 'Daman & Diu', '26': 'Dadra & Nagar Haveli and Daman & Diu',
    '27': 'Maharashtra', '28': 'Andhra Pradesh', '29': 'Karnataka',
    '30': 'Goa', '31': 'Lakshadweep', '32': 'Kerala',
    '33': 'Tamil Nadu', '34': 'Puducherry', '35': 'Andaman & Nicobar Islands',
    '36': 'Telangana', '37': 'Andhra Pradesh', '38': 'Ladakh',
    '97': 'Other Territory',
}


def state_from_gstin(gstin):
    gstin = str(gstin or '').strip().upper()
    code = gstin[:2] if len(gstin) >= 2 else ''
    return STATE_CODES.get(code, f'Unknown / Code {code}' if code else 'Unknown State')


def copy_row_style_all(ws, source_row, target_row):
    copy_row_style(ws, source_row, target_row)


def style_section_row(ws, row, text, fill_color='FFEFE2E7'):
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
        cell.font = Font(name='Trebuchet MS', size=10, bold=True, color='FF98002E')
        cell.alignment = Alignment(vertical='center')
        cell.border = Border(bottom=Side(style='thin', color='FFD0B5BF'))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    ws.cell(row, 1).value = text
    ws.cell(row, 1).font = Font(name='Trebuchet MS', size=11, bold=True, color='FF98002E')
    ws.cell(row, 1).alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 23


def add_total_style(ws, row, label='Total'):
    total_font = Font(name='Trebuchet MS', size=10, bold=True, color='FF98002E')
    total_fill = PatternFill(fill_type='solid', fgColor='FFF2E2E6')
    top = Side(style='double', color='FF98002E')
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row, c)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = Border(top=top, bottom=cell.border.bottom, left=cell.border.left, right=cell.border.right)
    ws.cell(row, 1).value = label
    ws.cell(row, 2).value = None


def normalize_workbook_fonts(wb):
    """Apply the requested workbook-wide font standard without changing styling.

    Every cell in every worksheet uses Trebuchet MS at exactly 10 pt. Existing
    bold/italic/underline/color/etc. attributes are preserved.
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font:
                    cell.font = cell.font.copy(name="Trebuchet MS", size=10)


def build_workbook(records, template_path):
    wb, ws, groups, month_col = load_template(template_path)

    # Clear old data rows but preserve template headers/formatting.
    start_row = 5
    old_max_row = ws.max_row
    for r in range(start_row, old_max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    # Keep the original template columns exactly where they are.
    # Column A is used for GSTIN/registration, B remains Month, and the
    # Filing Status column is appended after the existing GSTR-1 groups.
    # Do NOT insert a column here: openpyxl does not reliably shift merged
    # header ranges when insert_cols() is used, which breaks the template
    # header alignment.
    ws.cell(3, 1).value = 'GSTIN / Registration'
    ws.cell(4, 1).value = 'GSTIN'
    ws.cell(3, 1).font = Font(name='Trebuchet MS', size=10, bold=True, color='FFFFFFFF')
    ws.cell(4, 1).font = Font(name='Trebuchet MS', size=10, bold=True, color='FFFFFFFF')
    ws.column_dimensions['A'].width = 24

    # Append status after the final existing GSTR-1 section so every original
    # template header/group remains perfectly aligned.
    status_col = max(ws.max_column, max((g['end'] for g in groups), default=month_col)) + 1
    ws.cell(3, status_col).value = 'Filing Status'
    ws.cell(4, status_col).value = 'Status'
    for rr in (3, 4):
        cell = ws.cell(rr, status_col)
        cell.font = Font(name='Trebuchet MS', size=10, bold=True, color='FFFFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='FF98002E')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(status_col)].width = 16

    enriched = []
    for rec in records:
        gstin = rec.get('gstin') or 'GSTIN Not Found'
        enriched.append({**rec, 'gstin_display': gstin, 'state': state_from_gstin(rec.get('gstin'))})

    state_order = {}
    for rec in enriched:
        state_order.setdefault(rec['state'], []).append(rec)
    state_order = dict(sorted(state_order.items(), key=lambda kv: kv[0]))

    row = start_row
    audit_rows = []
    data_rows_by_state = {}

    # Helper for numeric formatting.
    number_fmt = '#,##0.00;[RED]-#,##0.00'
    count_fmt = '#,##0'

    for state, state_records in state_order.items():
        # State separator row.
        if row > start_row:
            row += 1
        state_header_row = row
        copy_row_style(ws, start_row, row)
        style_section_row(ws, row, f'{state}  •  {len({r["gstin_display"] for r in state_records})} registration(s)')
        row += 1

        # Keep each registration distinct inside the state block.
        gstin_groups = {}
        for rec in state_records:
            gstin_groups.setdefault(rec['gstin_display'], []).append(rec)

        state_data_start = row
        for gstin, gstin_records in sorted(gstin_groups.items(), key=lambda kv: kv[0]):
            for rec in sorted(gstin_records, key=lambda x: x['month'] or datetime.max):
                if row > start_row:
                    copy_row_style(ws, start_row, row)
                ws.cell(row, 1).value = gstin
                ws.cell(row, 2).value = rec.get('month')
                ws.cell(row, month_col).number_format = 'mmm-yy'
                ws.cell(row, status_col).value = rec.get('filing_status', 'FILED')

                for g in groups:
                    d = rec['sections'].get(g['title'], zero())
                    write_group_row(ws, row, g, d)
                    for i, header in enumerate(g['headers']):
                        c = g['start'] + i
                        h = header.strip().lower()
                        if isinstance(ws.cell(row, c).value, (int, float)):
                            ws.cell(row, c).number_format = count_fmt if h.startswith('no. of records') else number_fmt

                for group, source_line, heading in rec.get('audit', []):
                    audit_rows.append([
                        rec.get('file_name', ''), rec.get('gstin', ''), rec.get('state', ''),
                        rec.get('tax_period', ''), group, heading, source_line,
                    ])
                row += 1

        state_data_end = row - 1
        data_rows_by_state[state] = (state_data_start, state_data_end)

        # State total row.
        state_total_row = row
        if state_data_end >= state_data_start:
            copy_row_style(ws, start_row, row)
            for c in range(1, ws.max_column + 1):
                h = (ws.cell(4, c).value or '').strip().lower()
                if c == 1:
                    ws.cell(row, c).value = f'{state} Total'
                elif c in (month_col, status_col):
                    ws.cell(row, c).value = None
                else:
                    vals = [ws.cell(rr, c).value for rr in range(state_data_start, state_data_end + 1)]
                    total = sum(v for v in vals if isinstance(v, (int, float)))
                    ws.cell(row, c).value = int(total) if h.startswith('no. of records') else total
                    ws.cell(row, c).number_format = count_fmt if h.startswith('no. of records') else number_fmt
            add_total_style(ws, row, f'{state} Total')
        row += 1

    # Grand total across every state/registration.
    total_row = row
    if records:
        copy_row_style(ws, start_row, total_row)
    ws.cell(total_row, 1).value = 'GRAND TOTAL – ALL STATES / REGISTRATIONS'
    ws.cell(total_row, month_col).value = None
    ws.cell(total_row, status_col).value = None

    # Sum only the actual data rows (not state headers or state totals) to avoid double counting.
    all_data_rows = []
    for state, (s, e) in data_rows_by_state.items():
        all_data_rows.extend(range(s, e + 1))
    for g in groups:
        for i, header in enumerate(g['headers']):
            c = g['start'] + i
            h = header.strip().lower()
            total = 0.0
            for rr in all_data_rows:
                v = ws.cell(rr, c).value
                if isinstance(v, (int, float)):
                    total += v
            ws.cell(total_row, c).value = int(total) if h.startswith('no. of records') else total
            ws.cell(total_row, c).number_format = count_fmt if h.startswith('no. of records') else number_fmt
    add_total_style(ws, total_row, 'GRAND TOTAL – ALL STATES / REGISTRATIONS')

    # Format all group columns and keep the template's widths readable.
    for g in groups:
        width = 16 if any('value' in h.lower() for h in g['headers']) else 13
        for c in range(g['start'], g['end'] + 1):
            ws.column_dimensions[get_column_letter(c)].width = max(
                ws.column_dimensions[get_column_letter(c)].width or 0, width
            )
    ws.column_dimensions[get_column_letter(month_col)].width = 12

    if old_max_row > total_row:
        ws.delete_rows(total_row + 1, old_max_row - total_row)

    ws.freeze_panes = f'{get_column_letter(month_col + 1)}{start_row}'
    ws.auto_filter.ref = (
        f'A4:{get_column_letter(ws.max_column)}{total_row}'
    )

    # Extraction Audit.
    if 'Extraction Audit' in wb.sheetnames:
        del wb['Extraction Audit']
    audit = wb.create_sheet('Extraction Audit')
    headers = ['PDF File', 'GSTIN', 'State', 'Tax Period', 'Excel Group', 'GSTR-1 Section', 'Source Row']
    for c, h in enumerate(headers, start=1):
        audit.cell(1, c).value = h
        audit.cell(1, c).font = Font(bold=True)
    for r_idx, values in enumerate(audit_rows, start=2):
        for c_idx, value in enumerate(values, start=1):
            audit.cell(r_idx, c_idx).value = value
    audit.freeze_panes = 'A2'
    audit.auto_filter.ref = f'A1:G{max(1, len(audit_rows)+1)}'
    for col, width in {'A':35, 'B':20, 'C':28, 'D':15, 'E':45, 'F':70, 'G':90}.items():
        audit.column_dimensions[col].width = width

    # Filing Status sheet: one row per uploaded filing, including NIL months.
    if 'Filing Status' in wb.sheetnames:
        del wb['Filing Status']
    filing = wb.create_sheet('Filing Status')
    filing_headers = ['GSTIN', 'State', 'Month', 'Tax Period', 'Financial Year', 'Filing Status', 'PDF File']
    for c, h in enumerate(filing_headers, start=1):
        filing.cell(1, c).value = h
        filing.cell(1, c).font = Font(bold=True)
    for r_idx, rec in enumerate(sorted(records, key=lambda x: ((x.get('gstin') or ''), x.get('month') or datetime.max)), start=2):
        values = [
            rec.get('gstin'),
            rec.get('state') or state_from_gstin(rec.get('gstin')),
            rec.get('month'),
            rec.get('tax_period'),
            rec.get('fy'),
            rec.get('filing_status', 'FILED'),
            Path(rec.get('file_name', '')).name,
        ]
        for c, value in enumerate(values, start=1):
            filing.cell(r_idx, c).value = value
        filing.cell(r_idx, 3).number_format = 'mmm-yy'
    filing.freeze_panes = 'A2'
    filing.auto_filter.ref = f'A1:G{max(1, len(records)+1)}'
    for col, width in {'A':20, 'B':25, 'C':12, 'D':15, 'E':15, 'F':16, 'G':45}.items():
        filing.column_dimensions[col].width = width

    # Read Me.
    if 'Read Me' in wb.sheetnames:
        del wb['Read Me']
    meta = wb.create_sheet('Read Me', 0)
    meta.sheet_view.showGridLines = False

    maroon = 'FF98002E'
    title_font = Font(name='Trebuchet MS', size=16, bold=True, color=maroon)
    subtitle_font = Font(name='Trebuchet MS', size=10, italic=True, color='FF666666')
    label_font = Font(name='Trebuchet MS', size=10, bold=True, color=maroon)
    body_font = Font(name='Trebuchet MS', size=10, color='FF333333')
    link_font = Font(name='Trebuchet MS', size=10, bold=True, color='FF1155CC', underline='single')

    meta['A1'] = 'GSTR-1 PDF to Excel Analyzer'
    meta['A1'].font = title_font
    meta['A2'] = 'Powered by pushpakkumar.com'
    meta['A2'].font = subtitle_font
    meta['A2'].hyperlink = 'https://pushpakkumar.com'

    gstins = sorted({r.get('gstin') for r in records if r.get('gstin')})
    states = sorted({state_from_gstin(r.get('gstin')) for r in records})
    gstin_label = 'GSTINs / Registrations'
    if gstins:
        gstin_value = f'{len(gstins)} registration(s): ' + ', '.join(gstins)
    else:
        gstin_value = 'Not found in the uploaded PDF(s)'

    rows = [
        ('States', f'{len(states)} state(s): ' + ', '.join(states)),
        ('Invoice Value rule', 'Taxable Value + IGST + CGST + SGST + Cess'),
        ('Source', 'Filed GSTR-1 PDFs uploaded in the ZIP'),
        ('Grouping', 'Main sheet is grouped State-wise, then GSTIN-wise, then Month-wise. NIL-filed months are retained with zero values and marked NIL FILED. Each state has its own subtotal followed by one Grand Total for all states/registrations.'),
        ('NIL filing rule', 'A NIL GSTR-1 PDF is treated as a valid filing. The month is included in the report, Filing Status is marked NIL FILED, and all mapped values are explicitly set to zero.'),
        ('Mapping rule', 'Each Excel group is mapped to a specific GSTR-1 section/row. The PDF Value field is treated as Taxable Value.'),
        ('Audit', 'Extraction Audit records the exact source row used for each mapped section, along with GSTIN and State.'),
        ('Reconciliation', 'The tool captures the filed Total Liability and the difference against extracted taxable-outward invoice value; it does not overwrite source data.'),
    ]
    r = 4
    meta.cell(r, 1, gstin_label).font = label_font
    meta.cell(r, 2, gstin_value).font = body_font
    meta.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')
    r += 1
    for label, body in rows:
        meta.cell(r, 1, label).font = label_font
        meta.cell(r, 2, body).font = body_font
        meta.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')
        r += 1

    r += 1
    meta.cell(r, 1, 'Website').font = label_font
    link_cell = meta.cell(r, 2, 'pushpakkumar.com')
    link_cell.font = link_font
    link_cell.hyperlink = 'https://pushpakkumar.com'
    meta.column_dimensions['A'].width = 24
    meta.column_dimensions['B'].width = 100
    for rr in range(4, r + 1):
        meta.row_dimensions[rr].height = 18

    ws['A1'] = 'pushpakkumar.com'
    ws['A1'].font = Font(name='Trebuchet MS', size=10, italic=True, color='FF98002E')
    ws['A1'].hyperlink = 'https://pushpakkumar.com'

    # Final workbook-wide typography pass: every sheet/cell is Trebuchet MS,
    # 10 pt. This also normalizes fonts inherited from the original template
    # and fonts created by the audit/status/read-me sheets above.
    normalize_workbook_fonts(wb)

    return wb


def main():
    import streamlit as st

    st.set_page_config(
        page_title="Pushpak Kumar | GSTR-1 Analyzer",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    # Professional Pushpak Kumar branding
    st.markdown("""
    <style>
    .stApp { background:#f7f8fb; }
    .main .block-container {
        max-width:1180px;
        padding-top:2rem;
        padding-bottom:3rem;
    }

    .hero {
        background:linear-gradient(135deg,#98002e 0%,#65001f 100%);
        border-radius:20px;
        padding:30px 34px;
        color:#fff;
        margin-bottom:22px;
        box-shadow:0 10px 30px rgba(90,0,30,.16);
    }

    .brand-row {
        display:flex;
        align-items:center;
        gap:12px;
        margin-bottom:20px;
    }

    .brand-logo {
        width:44px;
        height:44px;
        border-radius:12px;
        background:rgba(255,255,255,.14);
        border:1px solid rgba(255,255,255,.25);
        display:flex;
        align-items:center;
        justify-content:center;
        font-size:17px;
        font-weight:800;
        letter-spacing:-1px;
    }

    .brand-name {
        font-size:14px;
        font-weight:750;
        letter-spacing:.4px;
        color:#fff;
    }

    .brand-tagline {
        font-size:11px;
        color:rgba(255,255,255,.72);
        margin-top:2px;
    }

    .hero h1 {
        margin:0 0 8px 0;
        font-size:31px;
        font-weight:750;
        letter-spacing:-.6px;
    }

    .hero p {
        margin:0;
        color:rgba(255,255,255,.88);
        font-size:14px;
    }

    .card {
        background:#fff;
        border:1px solid #e7e9ef;
        border-radius:16px;
        padding:22px;
        margin-bottom:18px;
        box-shadow:0 3px 14px rgba(20,20,40,.05);
    }

    .section-title {
        font-size:18px;
        font-weight:700;
        color:#242733;
        margin-bottom:4px;
    }

    .section-subtitle {
        color:#707582;
        font-size:13px;
        margin-bottom:14px;
    }

    .workflow {
        display:flex;
        gap:10px;
        align-items:center;
        color:#424752;
        font-size:13px;
    }

    .step {
        display:flex;
        align-items:center;
        gap:8px;
        flex:1;
    }

    .step-num {
        width:28px;
        height:28px;
        border-radius:50%;
        background:#f3e6eb;
        color:#98002e;
        display:flex;
        align-items:center;
        justify-content:center;
        font-weight:750;
    }

    .workflow-line {
        height:1px;
        background:#dddfe6;
        flex:1;
    }

    .status-card {
        background:#fff;
        border-left:4px solid #98002e;
        border-radius:10px;
        padding:13px 16px;
        margin:12px 0 16px;
        box-shadow:0 2px 9px rgba(20,20,40,.04);
    }

    .brand-chip {
        display:inline-flex;
        align-items:center;
        gap:7px;
        background:#f3e6eb;
        color:#98002e;
        border-radius:999px;
        padding:6px 11px;
        font-size:11px;
        font-weight:750;
        margin-bottom:11px;
    }

    .brand-dot {
        width:6px;
        height:6px;
        border-radius:50%;
        background:#98002e;
    }

    .footer {
        text-align:center;
        color:#858994;
        font-size:12px;
        padding-top:22px;
    }

    .footer-name {
        color:#98002e;
        font-weight:750;
    }

    div[data-testid="stFileUploader"] {
        background:#fbfbfd;
        border:1px dashed #c9ccd6;
        border-radius:13px;
        padding:8px;
    }

    .stButton > button,
    .stDownloadButton > button {
        border-radius:10px;
        min-height:44px;
        font-weight:650;
    }
    </style>
    """, unsafe_allow_html=True)

    # Header
    st.markdown("""
    <div class="hero">
        <div class="brand-row">
            <div class="brand-logo">PK</div>
            <div>
                <div class="brand-name">PUSHPAK KUMAR</div>
                <div class="brand-tagline">GSTR-1 Analysis &amp; Reporting</div>
            </div>
        </div>
        <h1>GSTR-1 PDF → Excel Analyzer</h1>
        <p>Automate filed GSTR-1 analysis, section mapping and Excel reporting.</p>
    </div>
    """, unsafe_allow_html=True)

    # Workflow
    st.markdown("""
    <div class="card">
        <div class="workflow">
            <div class="step"><span class="step-num">1</span><b>Upload ZIP</b></div>
            <div class="workflow-line"></div>
            <div class="step"><span class="step-num">2</span><b>Analyze PDFs</b></div>
            <div class="workflow-line"></div>
            <div class="step"><span class="step-num">3</span><b>Download Excel</b></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Upload
    st.markdown("""
    <div class="card">
        <div class="brand-chip"><span class="brand-dot"></span> PUSHPAK KUMAR • ANALYZER</div>
        <div class="section-title">Upload GSTR-1 files</div>
        <div class="section-subtitle">
            Upload a ZIP containing filed GSTR-1 PDF copies from one or more GST registrations. The report is grouped state-wise, then registration-wise, then month-wise.
        </div>
    """, unsafe_allow_html=True)

    template = Path(__file__).with_name("Sample_format.xlsx")
    zip_file = st.file_uploader(
        "Upload ZIP containing GSTR-1 filed PDFs",
        type=["zip"],
        label_visibility="collapsed",
    )

    st.markdown("</div>", unsafe_allow_html=True)

    if not zip_file:
        st.info("Tip: You can upload multiple registrations in one ZIP. The Excel report will group them State-wise and provide state subtotals plus one Grand Total.")
        st.markdown("""
        <div class="footer">
            Invoice Value = Taxable Value + IGST + CGST + SGST + Cess
            <br><br>
            Designed &amp; developed by <span class="footer-name">Pushpak Kumar</span>
        </div>
        """, unsafe_allow_html=True)
        return

    # Inspect ZIP before processing
    try:
        with zipfile.ZipFile(zip_file) as z_preview:
            pdf_names = [
                n for n in z_preview.namelist()
                if n.lower().endswith(".pdf") and not n.endswith("/")
            ]
    except zipfile.BadZipFile:
        st.error("The uploaded file is not a valid ZIP file.")
        return

    if not pdf_names:
        st.error("No PDF files were found in the ZIP.")
        return

    size_mb = len(zip_file.getvalue()) / (1024 * 1024)

    st.markdown(
        f"""
        <div class="status-card">
            <b>Ready to analyze</b><br>
            <span style="color:#707582;font-size:13px;">
                {len(pdf_names)} PDF file(s) &nbsp; • &nbsp; {size_mb:.2f} MB
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if st.button("🔍  Analyze GSTR-1 ZIP", type="primary", use_container_width=True):
        records = []
        errors = []

        progress = st.progress(0, text="Starting analysis…")
        status = st.empty()

        total_files = len(pdf_names)

        with zipfile.ZipFile(zip_file) as z:
            for i, name in enumerate(pdf_names, start=1):
                status.markdown(
                    f"**Processing {i} of {total_files}:** `{Path(name).name}`"
                )

                try:
                    d = extract_pdf(z.read(name))

                    # Some downloaded NIL copies are named with NIL in the
                    # filename. Treat that as an additional explicit marker,
                    # but never infer NIL merely because the PDF contains the
                    # word "Nil" (Table 8 routinely does).
                    if re.search(r'(^|[_\- .])NIL([_\- .]|$)', Path(name).stem, re.I):
                        d['is_nil'] = True
                        d['filing_status'] = 'NIL FILED'
                        for key in list(d.get('sections', {})):
                            d['sections'][key] = zero()
                        d['total_liability'] = 0.0
                        d['reconciliation'] = {
                            'total_liability': 0.0,
                            '6A_invoice_value': 0.0,
                            'difference': 0.0,
                        }

                    if not d.get("month"):
                        raise ValueError(
                            "Could not identify Financial Year / Tax Period"
                        )

                    d["file_name"] = name
                    records.append(d)

                except Exception as exc:
                    errors.append((name, str(exc)))

                progress.progress(
                    i / total_files,
                    text=f"Analyzing PDF {i} of {total_files}",
                )

        progress.progress(1.0, text="Analysis complete ✓")
        status.empty()

        if errors:
            st.warning(f"{len(errors)} file(s) could not be parsed.")
            with st.expander("View parsing errors"):
                st.dataframe(
                    [{"File": n, "Error": e} for n, e in errors],
                    use_container_width=True,
                    hide_index=True,
                )

        if not records:
            st.error("No GSTR-1 PDFs could be successfully parsed.")
            return

        st.success(
            f"Successfully parsed {len(records)} of {len(pdf_names)} PDF(s)."
        )

        # Summary cards
        gstins = sorted(
            {r.get("gstin") for r in records if r.get("gstin")}
        )
        fys = sorted(
            {r.get("fy") for r in records if r.get("fy")}
        )

        a, b, c = st.columns(3)
        with a:
            st.metric("PDFs Parsed", len(records))
        with b:
            st.metric("GSTINs Found", len(gstins))
        with c:
            st.metric("Financial Year(s)", len(fys))

        if len(gstins) > 1:
            st.info(
                "Multiple GSTINs detected. The Excel report will automatically group them State-wise, then GSTIN-wise, with state subtotals and one Grand Total."
            )

        # Period preview
        st.markdown(
            """
            <div class="section-title" style="margin-top:22px;">
                Extracted periods
            </div>
            <div class="section-subtitle">
                Review the detected tax periods before downloading.
            </div>
            """,
            unsafe_allow_html=True,
        )

        preview = []
        for r in sorted(
            records,
            key=lambda x: x["month"] or datetime.max
        ):
            preview.append({
                "Month": r["month"].strftime("%b-%y")
                    if r.get("month") else "",
                "GSTIN": r.get("gstin"),
                "Tax Period": r.get("tax_period"),
                "Financial Year": r.get("fy"),
                "Status": r.get("filing_status", "FILED"),
                "File": Path(r["file_name"]).name,
            })

        st.dataframe(
            preview,
            use_container_width=True,
            hide_index=True,
        )

        # Excel build progress
        build_progress = st.progress(
            0,
            text="Preparing Excel workbook…"
        )

        build_progress.progress(
            25,
            text="Loading Excel template…"
        )

        build_progress.progress(
            50,
            text="Applying GSTR-1 mappings…"
        )

        wb = build_workbook(records, str(template))

        build_progress.progress(
            75,
            text="Creating extraction audit trail…"
        )

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        build_progress.progress(
            100,
            text="Excel report ready ✓"
        )

        st.markdown(
            """
            <div class="card">
                <div class="section-title">Your report is ready</div>
                <div class="section-subtitle">
                    The workbook includes the mapped report, Extraction Audit
                    and Read Me sheets.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.download_button(
            "⬇️  Download GSTR1_Analyzed.xlsx",
            out.getvalue(),
            file_name="GSTR1_Analyzed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    st.markdown("""
    <div class="footer">
        GSTR-1 PDF → Excel Analyzer
        &nbsp; • &nbsp;
        Invoice Value = Taxable Value + IGST + CGST + SGST + Cess
        <br><br>
        Designed &amp; developed by
        <span class="footer-name">Pushpak Kumar</span>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
